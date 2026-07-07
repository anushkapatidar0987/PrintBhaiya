import csv
import logging
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.db.models import Count, Sum, Q
from django.utils import timezone
from django.core.mail import EmailMessage
from django.conf import settings
from rest_framework import generics, permissions, status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser

from core.permissions import IsSuperAdmin
from core.models import BaseModel
from shops.models import Shop, ShopStatusLog
from orders.models import Order
from accounts.models import User
from shops.serializers import ShopListSerializer
from orders.serializers import OrderListSerializer

from .models import Notice, SupportTicket
from .serializers import NoticeSerializer, SupportTicketSerializer, ShopStatusLogSerializer

logger = logging.getLogger(__name__)

class PendingShopsView(generics.ListAPIView):
    """List all shops pending approval."""
    permission_classes = [IsSuperAdmin]
    serializer_class = ShopListSerializer

    def get_queryset(self):
        return Shop.objects.filter(is_approved=False)


class SuperAdminShopListView(generics.ListAPIView):
    """List all registered shops for the superadmin (both active and hidden/inactive)."""
    permission_classes = [IsSuperAdmin]
    serializer_class = ShopListSerializer
    queryset = Shop.objects.all().order_by('-created_at')


class ApproveShopView(APIView):
    """Approve a pending shop."""
    permission_classes = [IsSuperAdmin]

    def patch(self, request, id):
        shop = get_object_or_404(Shop, id=id)
        if shop.is_approved:
            return Response({'message': 'Shop is already approved'})
            
        shop.is_approved = True
        shop.save()
        
        # Write initial status log
        ShopStatusLog.objects.create(
            shop=shop,
            status=shop.status,
            changed_by=request.user
        )
        return Response({'message': f'Shop {shop.name} approved successfully'})


class PlatformAnalyticsView(APIView):
    """Basic platform analytics for the dashboard."""
    permission_classes = [IsSuperAdmin]

    def get(self, request):
        total_students = User.objects.filter(role=User.Role.STUDENT).count()
        total_shops = Shop.objects.filter(is_approved=True).count()
        
        order_stats = Order.objects.aggregate(
            total_orders=Count('id'),
            total_revenue=Sum('total_amount'),
            total_platform_fee=Sum('platform_fee')
        )
        
        pending_shops = Shop.objects.filter(is_approved=False).count()
        
        return Response({
            'users': {
                'total_students': total_students,
                'approved_shops': total_shops,
                'pending_shop_approvals': pending_shops
            },
            'orders': {
                'total_orders': order_stats['total_orders'] or 0,
                'total_revenue': float(order_stats['total_revenue'] or 0),
                'total_platform_fee': float(order_stats['total_platform_fee'] or 0)
            }
        })


class AllOrdersView(generics.ListAPIView):
    """List all platform orders for the admin."""
    permission_classes = [IsSuperAdmin]
    serializer_class = OrderListSerializer
    queryset = Order.objects.all().order_by('-created_at')


# --- MODULE 1: Notices & Broadcasts ---

class NoticeListCreateView(generics.ListCreateAPIView):
    """Admin-only view to create global or targeted announcements."""
    permission_classes = [IsSuperAdmin]
    serializer_class = NoticeSerializer
    queryset = Notice.objects.all().order_by('-created_at')


class StudentNoticeListView(generics.ListAPIView):
    """Endpoint for students to fetch relevant notifications/broadcasts."""
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = NoticeSerializer

    def get_queryset(self):
        user = self.request.user
        if user.role == User.Role.STUDENT:
            return Notice.objects.filter(
                Q(audience_type=Notice.Audience.ALL) |
                Q(audience_type=Notice.Audience.STUDENTS) |
                Q(recipients=user)
            ).distinct().order_by('-created_at')
        elif user.role == User.Role.SHOP_OWNER:
            return Notice.objects.filter(
                Q(audience_type=Notice.Audience.ALL) |
                Q(audience_type=Notice.Audience.SHOPKEEPERS) |
                Q(recipients=user)
            ).distinct().order_by('-created_at')
        return Notice.objects.all().order_by('-created_at')


# --- MODULE 2: Shop History, Status & Toggle ---

class ShopStatusHistoryView(generics.ListAPIView):
    """View to fetch status logs and analytics of a specific shop."""
    permission_classes = [IsSuperAdmin]
    serializer_class = ShopStatusLogSerializer

    def get_queryset(self):
        shop_id = self.kwargs.get('id')
        return ShopStatusLog.objects.filter(shop_id=shop_id).order_by('-created_at')


class ToggleShopActiveView(APIView):
    """Admin soft-toggle to hide/unhide a shop from the student portal."""
    permission_classes = [IsSuperAdmin]

    def patch(self, request, id):
        shop = get_object_or_404(Shop, id=id)
        shop.is_active = not shop.is_active
        shop.save()
        status_str = "visible" if shop.is_active else "hidden"
        return Response({
            'message': f'Shop {shop.name} is now {status_str} on the listing platform.',
            'is_active': shop.is_active
        })


# --- MODULE 4: CSV Export & Filters ---

class ExportUsersCSVView(APIView):
    """Compile and export user registration logs with advanced filtering options (CSV, Excel, PDF)."""
    permission_classes = [IsSuperAdmin]

    def get(self, request):
        role_filter = request.query_params.get('role')  # STUDENT, SHOP_OWNER
        status_filter = request.query_params.get('status')  # active, inactive
        date_start = request.query_params.get('date_start')
        date_end = request.query_params.get('date_end')
        campus_wing = request.query_params.get('campus_wing')
        min_orders = request.query_params.get('min_orders')
        export_format = request.query_params.get('format', 'csv').lower()

        queryset = User.objects.all()

        if role_filter:
            queryset = queryset.filter(role=role_filter)
        if status_filter:
            is_active = status_filter.lower() == 'active'
            queryset = queryset.filter(is_active=is_active)
        if date_start:
            queryset = queryset.filter(date_joined__date__gte=date_start)
        if date_end:
            queryset = queryset.filter(date_joined__date__lte=date_end)
        if campus_wing:
            queryset = queryset.filter(campus_wing__icontains=campus_wing)

        queryset = queryset.annotate(order_count=Count('orders'))
        if min_orders:
            try:
                queryset = queryset.filter(order_count__gte=int(min_orders))
            except ValueError:
                pass

        filename_prefix = f"users_report_{timezone.now().strftime('%Y%m%d_%H%M%S')}"

        # 1. EXCEL FORMAT
        if export_format == 'xlsx':
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "Users Registry"
            ws.views.sheetView[0].showGridLines = True

            # Colors & Styles
            navy_header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
            soft_gray_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
            font_title = Font(name="Calibri", size=16, bold=True, color="4F46E5")
            font_subtitle = Font(name="Calibri", size=10, italic=True, color="6B7280")
            font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            font_data = Font(name="Calibri", size=11, color="1F2937")
            font_bold = Font(name="Calibri", size=11, bold=True, color="1F2937")

            align_center = Alignment(horizontal="center", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center")
            align_right = Alignment(horizontal="right", vertical="center")

            thin_border_side = Side(border_style="thin", color="E5E7EB")
            thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

            # Document Title Header Block
            ws.merge_cells("A1:I1")
            ws["A1"] = "PrintKarDoBhaiya - Platform Registry Report"
            ws["A1"].font = font_title
            ws["A1"].alignment = align_left

            ws.merge_cells("A2:I2")
            ws["A2"] = f"Generated on {timezone.now().strftime('%Y-%m-%d %H:%M:%S')} | Target Audience: {role_filter or 'Global'}"
            ws["A2"].font = font_subtitle
            ws["A2"].alignment = align_left

            # Headers Row
            headers = ['ID', 'Name', 'Email', 'Phone Number', 'Role', 'Status', 'Campus Wing', 'Order Volume', 'Joined Date']
            ws.append([]) # row 3 spacing
            ws.append(headers) # row 4 headers

            for col_num in range(1, 10):
                cell = ws.cell(row=4, column=col_num)
                cell.font = font_header
                cell.fill = navy_header_fill
                cell.alignment = align_center
                cell.border = thin_border

            ws.row_dimensions[4].height = 25

            # Rows compilation
            row_idx = 5
            for user in queryset:
                ws.append([
                    str(user.id),
                    user.get_full_name() or 'N/A',
                    user.email,
                    user.phone_number,
                    user.get_role_display(),
                    'Active' if user.is_active else 'Inactive',
                    user.campus_wing or 'N/A',
                    user.order_count,
                    user.date_joined.strftime('%Y-%m-%d %H:%M:%S')
                ])
                
                fill = soft_gray_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)
                for col_num in range(1, 10):
                    cell = ws.cell(row=row_idx, column=col_num)
                    cell.font = font_data
                    cell.border = thin_border
                    if fill.fill_type:
                        cell.fill = fill
                    
                    if col_num in [1, 5, 6, 9]:
                        cell.alignment = align_center
                    elif col_num == 8:
                        cell.alignment = align_right
                        cell.font = font_bold
                    else:
                        cell.alignment = align_left
                        
                ws.row_dimensions[row_idx].height = 20
                row_idx += 1

            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename_prefix}.xlsx"'
            wb.save(response)
            return response

        # 2. PDF FORMAT
        elif export_format == 'pdf':
            from io import BytesIO
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

            buffer = BytesIO()
            doc = SimpleDocTemplate(
                buffer, 
                pagesize=landscape(letter), 
                rightMargin=25, 
                leftMargin=25, 
                topMargin=25, 
                bottomMargin=25
            )
            story = []
            styles = getSampleStyleSheet()

            # Styles Setup
            title_style = ParagraphStyle(
                'DocTitle',
                parent=styles['Heading1'],
                fontSize=20,
                textColor=colors.HexColor('#1E293B'),
                spaceAfter=4
            )
            subtitle_style = ParagraphStyle(
                'DocSubTitle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#64748B'),
                spaceAfter=15
            )
            cell_style = ParagraphStyle(
                'CellText',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.HexColor('#334155')
            )
            header_style = ParagraphStyle(
                'HeaderCellText',
                parent=styles['Normal'],
                fontSize=9,
                bold=True,
                textColor=colors.white
            )

            story.append(Paragraph("PrintKarDoBhaiya - Platform Registry Report", title_style))
            story.append(Paragraph(f"Generated on {timezone.now().strftime('%Y-%m-%d %H:%M:%S')} | Target: {role_filter or 'Global'} | Wing: {campus_wing or 'All'}", subtitle_style))
            story.append(Spacer(1, 8))

            col_widths = [60, 100, 140, 90, 70, 50, 70, 50, 102]
            table_data = []

            # Headers row
            table_data.append([
                Paragraph("ID", header_style),
                Paragraph("Name", header_style),
                Paragraph("Email", header_style),
                Paragraph("Phone", header_style),
                Paragraph("Role", header_style),
                Paragraph("Status", header_style),
                Paragraph("Campus Wing", header_style),
                Paragraph("Orders", header_style),
                Paragraph("Joined Date", header_style),
            ])

            # Data rows
            for user in queryset:
                short_uuid = str(user.id)[:8] + '...'
                table_data.append([
                    Paragraph(short_uuid, cell_style),
                    Paragraph(user.get_full_name() or 'N/A', cell_style),
                    Paragraph(user.email, cell_style),
                    Paragraph(user.phone_number, cell_style),
                    Paragraph(user.get_role_display(), cell_style),
                    Paragraph('Active' if user.is_active else 'Inactive', cell_style),
                    Paragraph(user.campus_wing or 'N/A', cell_style),
                    Paragraph(str(user.order_count), cell_style),
                    Paragraph(user.date_joined.strftime('%Y-%m-%d %H:%M'), cell_style),
                ])

            user_table = Table(table_data, colWidths=col_widths, repeatRows=1)
            t_style = TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('BOTTOMPADDING', (0,0), (-1,0), 6),
                ('TOPPADDING', (0,0), (-1,0), 6),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
            ])

            for i in range(1, len(table_data)):
                if i % 2 == 0:
                    t_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F8FAFC'))

            user_table.setStyle(t_style)
            story.append(user_table)

            doc.build(story)
            pdf_data = buffer.getvalue()
            buffer.close()

            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename_prefix}.pdf"'
            response.write(pdf_data)
            return response

        # 3. CSV FORMAT (DEFAULT)
        else:
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{filename_prefix}.csv"'
            writer = csv.writer(response)
            writer.writerow(['ID', 'Name', 'Email', 'Phone Number', 'Role', 'Status', 'Campus Wing', 'Order Volume', 'Joined Date'])

            for user in queryset:
                writer.writerow([
                    user.id,
                    user.get_full_name() or 'N/A',
                    user.email,
                    user.phone_number,
                    user.get_role_display(),
                    'Active' if user.is_active else 'Inactive',
                    user.campus_wing or 'N/A',
                    user.order_count,
                    user.date_joined.strftime('%Y-%m-%d %H:%M:%S')
                ])

            return response


# --- MODULE 5: Student Support Portal ---

from accounts.serializers import UserSerializer

class UsersListView(generics.ListAPIView):
    """List all registered users on the platform for admin review."""
    permission_classes = [IsSuperAdmin]
    serializer_class = UserSerializer
    queryset = User.objects.all().order_by('-date_joined')


class CreateSupportTicketView(APIView):
    """Endpoint for students to raise support issues with files."""
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        serializer = SupportTicketSerializer(data=request.data)
        if serializer.is_valid():
            ticket = serializer.save(student=request.user)

            # Trigger dynamic email alert
            try:
                self.send_support_email(ticket)
            except Exception as e:
                logger.error(f"Support ticket email failed to send: {e}")

            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def send_support_email(self, ticket):
        subject = f"[SUPPORT TICKET #{ticket.id[:8]}] - {ticket.get_category_display()}"
        body = f"""
New support ticket raised by: {ticket.student.get_full_name()} ({ticket.student.email})

Category: {ticket.get_category_display()}
Student Email: {ticket.email}
Student Phone: {ticket.phone_number}

Details:
{ticket.details}

-- 
PrintKarDoBhaiya Support Automation
"""
        email = EmailMessage(
            subject=subject,
            body=body,
            from_email=settings.DEFAULT_FROM_EMAIL if hasattr(settings, 'DEFAULT_FROM_EMAIL') else 'noreply@printkardobhaiya.com',
            to=['support@printkardobhaiya.com'],
        )

        if ticket.attachment:
            email.attach(ticket.attachment.name, ticket.attachment.read())

        email.send(fail_silently=False)
